from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.serializers.json import DjangoJSONEncoder
from django.db import transaction
from django.db.models import Sum, Q, Case, When, DecimalField, Value, F
from django.db.models.functions import TruncDate
from django.http import HttpResponse, JsonResponse, Http404, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.templatetags.static import static
from django.utils import timezone
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.urls import reverse
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import plotly.graph_objs as go
import plotly.offline as opy
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from simple_history.utils import update_change_reason
from .models import Venta, DetalleVenta, MovimientoCaja, Caja, procesar_venta
from .forms import VentaForm, DetalleVentaFormSet
from productos.models import Producto
from .models import revertir_venta

def anular_venta(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    if request.method == "POST":
        motivo = request.POST.get("motivo", "").strip()
        try:
            revertir_venta(venta, request.user, motivo)
            messages.success(request, "Venta anulada correctamente.")
        except Exception as e:
            messages.error(request, f"Error al anular la venta: {str(e)}")
    return redirect('ventas:lista_ventas')

def get_total_movimientos(caja):
    return caja.movimientos.aggregate(
        total=Sum(
            Case(
                When(tipo='INGRESO', then='monto'),
                When(tipo='EGRESO', then=-'monto'),
                output_field=DecimalField(),
                default=Value(0)
            )
        )
    )["total"] or Decimal("0.00")
    
def grupo_requerido(*nombres_grupos):
    def decorador(view_func):
        @login_required
        def _wrapped_view(request, *args, **kwargs):
            if request.user.groups.filter(name__in=nombres_grupos).exists():
                return view_func(request, *args, **kwargs)
            messages.error(request, "No tienes permisos para acceder a esta sección.")
            return redirect('reportes:dashboard')  # Ajusta si tu ruta de dashboard es distinta
        return _wrapped_view
    return decorador

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def ticket_venta_data(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    detalles = venta.detalles.select_related('producto')
    # Mapear método de pago al texto legible
    metodo_pago_texto = dict(Venta.METODOS_PAGO).get(venta.metodo_pago, venta.metodo_pago)
    if venta.metodo_pago == 'OT':
        metodo_pago_texto = f"Mixto (Efectivo: ${venta.mixto_efectivo or 0:.2f}, Otros: ${venta.mixto_otros or 0:.2f})"
    # Preparar datos para el ticket con fechas localizadas a TIME_ZONE (America/Cancun)
    data = {
        'venta': {
            'id': venta.id,
            'fecha': venta.fecha.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M'),
            'usuario': venta.usuario.get_full_name() or venta.usuario.username,
            'metodo_pago': metodo_pago_texto,
            'total': float(venta.total),
        },
        'detalles': [
            {
                'producto_nombre': detalle.producto.nombre,
                'cantidad': detalle.cantidad,
                'producto_precio': float(detalle.producto.precio),
                'subtotal': float(detalle.subtotal),
            } for detalle in detalles
        ],
        'hoy': timezone.now().astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M'),
    }
    return JsonResponse(data)

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def abrir_caja(request):
    if request.method == "POST":
        monto = request.POST.get("monto")
        try:
            monto = Decimal(monto)
        except:
            messages.error(request, "Monto inválido.")
            return redirect("ventas:abrir_caja")

        existe_abierta = Caja.objects.filter(usuario=request.user, esta_abierta=True).exists()
        if existe_abierta:
            messages.error(request, "Ya tienes una caja abierta.")
            return redirect("ventas:estado_caja_actual")

        Caja.objects.create(
            usuario=request.user,
            saldo_inicial=monto,
            esta_abierta=True,
        )
        messages.success(request, "Caja abierta correctamente.")
        return redirect("ventas:estado_caja_actual")
    return render(request, "ventas/caja/abrir_caja.html")

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def cerrar_caja(request):
    caja = Caja.objects.filter(usuario=request.user, esta_abierta=True).first()
    if not caja:
        messages.error(request, "No tienes una caja abierta.")
        return redirect("ventas:estado_caja_actual")
    if request.method == "POST":
        total_movimientos = caja.movimientos.aggregate(
            total=Sum(
                Case(
                    When(tipo='INGRESO', then=F('monto')),
                    When(tipo='EGRESO', then=-F('monto')),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                    default=Value(Decimal('0.00'))
                )
            )
        )["total"] or Decimal("0.00")
        caja.saldo_final = caja.saldo_inicial + total_movimientos
        caja.esta_abierta = False
        caja.fecha_cierre = now()
        caja.save()
        messages.success(request, f"Caja cerrada correctamente. Saldo final: ${caja.saldo_final:.2f}")
        return redirect("ventas:estado_caja_actual")
    return render(request, "ventas/caja/cerrar_caja.html", {"caja": caja})

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def estado_caja_actual(request):
    caja = Caja.objects.filter(usuario=request.user, esta_abierta=True).first()
    if not caja:
        messages.info(request, "No tienes caja abierta.")
        return render(request, "ventas/caja/estado_caja.html", {"caja": None})
    
    total_movimientos = caja.movimientos.aggregate(
        total=Sum(
            Case(
                When(tipo='INGRESO', then=F('monto')),
                When(tipo='EGRESO', then=-F('monto')),
                output_field=DecimalField(max_digits=10, decimal_places=2),
                default=Value(Decimal('0.00'))
            )
        )
    )["total"] or Decimal("0.00")
    saldo_actual = caja.saldo_inicial + total_movimientos
    return render(request, "ventas/caja/estado_caja.html", {
        "caja": caja,
        "saldo_actual": saldo_actual,
        "movimientos": caja.movimientos.order_by('-fecha'),
    })

@grupo_requerido('Encargado', 'Gerente')
def movimientos_caja(request):
    cajas = Caja.objects.all().order_by('-fecha_apertura')
    return render(request, "ventas/caja/movimientos.html", {"cajas": cajas})

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
@login_required
def lista_ventas(request):
    usuario = request.user
    estado = request.GET.get('estado', '').strip()
    usuario_id = request.GET.get('usuario', '').strip()
    # Guardar filtros en sesión
    if 'estado' in request.GET:
        request.session['ventas_estado_filtro'] = estado
    else:
        estado = request.session.get('ventas_estado_filtro', '')

    if 'usuario' in request.GET:
        request.session['ventas_usuario_filtro'] = usuario_id
    else:
        usuario_id = request.session.get('ventas_usuario_filtro', '')
    # Filtro base por estado
    if estado == 'activas':
        ventas = Venta.objects.filter(anulada=False)
    elif estado == 'anuladas':
        ventas = Venta.objects.filter(anulada=True)
    else:
        ventas = Venta.objects.all()
    # Control de acceso y filtrado por rol
    es_cajero = usuario.groups.filter(name='Cajero').exists()
    es_encargado_o_gerente = usuario.groups.filter(name__in=['Encargado', 'Gerente']).exists()
    if es_cajero:
        ventas = ventas.filter(usuario=usuario)
        usuario_id = str(usuario.id)  # Forzar valor fijo para mostrarlo en tabla si se usa
        usuarios_cajeros = User.objects.filter(id=usuario.id)  # solo él mismo
    elif es_encargado_o_gerente:
        if usuario_id:
            ventas = ventas.filter(usuario__id=usuario_id)
        usuarios_cajeros = User.objects.filter(groups__name='Cajero').order_by('username')
    else:
        # Bloqueo de seguridad (si se requiere)
        ventas = Venta.objects.none()
        usuarios_cajeros = User.objects.none()
    ventas = ventas.order_by('-fecha')
    context = {
        'ventas': ventas,
        'estado': estado,
        'usuario_id': usuario_id,
        'usuarios_cajeros': usuarios_cajeros,
        'es_gerente_o_encargado': es_encargado_o_gerente,
    }
    return render(request, 'ventas/lista_ventas.html', context)

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def crear_venta(request):
    if request.method == "POST":
        metodo_pago_texto = request.POST.get("metodo_pago")
        mapa_metodo = {
            "Efectivo": "EF",
            "Tarjeta": "TC",  # Assuming frontend uses "Tarjeta" for both TC and TD
            "Tarjeta Débito": "TD",
            "Transferencia": "TR",
            "Mixto": "OT",
        }
        metodo_pago = mapa_metodo.get(metodo_pago_texto)
        if not metodo_pago:
            messages.error(request, "Método de pago inválido.")
            return redirect("ventas:crear_venta")
        productos = []
        for key in request.POST:
            if key.startswith("productos[") and key.endswith("][id]"):
                index = key.split('[')[1].split(']')[0]
                prod_id = request.POST.get(f"productos[{index}][id]")
                cantidad = request.POST.get(f"productos[{index}][cantidad]")
                if prod_id and cantidad:
                    try:
                        productos.append({"id": int(prod_id), "cantidad": int(cantidad)})
                    except ValueError:
                        messages.error(request, "Cantidad o ID de producto inválidos.")
                        return redirect("ventas:crear_venta")
        if not productos:
            messages.error(request, "Debe agregar al menos un producto a la venta.")
            return redirect("ventas:crear_venta")
        total_venta = 0
        productos_objetos = []
        for p in productos:
            try:
                producto = Producto.objects.get(pk=p["id"])
                if producto.stock < p["cantidad"]:
                    messages.error(request, f"Stock insuficiente para {producto.nombre}.")
                    return redirect("ventas:crear_venta")
                productos_objetos.append((producto, p["cantidad"]))
                total_venta += producto.precio * p["cantidad"]
            except Producto.DoesNotExist:
                messages.error(request, "Producto no encontrado.")
                return redirect("ventas:crear_venta")
        # Validate payment method details
        mixto_efectivo = None
        mixto_otros = None
        if metodo_pago_texto == "Efectivo":
            try:
                monto_efectivo = Decimal(request.POST.get("monto_efectivo", "0"))
                if monto_efectivo < total_venta:
                    messages.error(request, "El monto entregado en efectivo es insuficiente.")
                    return redirect("ventas:crear_venta")
            except InvalidOperation:
                messages.error(request, "Monto en efectivo inválido.")
                return redirect("ventas:crear_venta")
        elif metodo_pago_texto in ["Tarjeta", "Tarjeta Débito", "Transferencia"]:
            if metodo_pago_texto == "Tarjeta":
                metodo_pago = "TC"  # Default to credit card if "Tarjeta" is selected
            numero_tarjeta = request.POST.get("numero_tarjeta", "").strip()
            if not numero_tarjeta or len(numero_tarjeta) < 4:
                messages.error(request, "Número de tarjeta o transferencia inválido.")
                return redirect("ventas:crear_venta")
        elif metodo_pago_texto == "Mixto":
            try:
                mixto_efectivo = Decimal(request.POST.get("mixto_efectivo", "0"))
                mixto_otros = Decimal(request.POST.get("mixto_tarjeta", "0"))
                if (mixto_efectivo + mixto_otros) < total_venta:
                    messages.error(request, "La suma de efectivo y otros es insuficiente.")
                    return redirect("ventas:crear_venta")
            except InvalidOperation:
                messages.error(request, "Montos mixtos inválidos.")
                return redirect("ventas:crear_venta")
        # Create sale
        venta = Venta.objects.create(
            usuario=request.user,
            metodo_pago=metodo_pago,
            total=total_venta,
            mixto_efectivo=mixto_efectivo if metodo_pago == "OT" else None,
            mixto_otros=mixto_otros if metodo_pago == "OT" else None,
        )
        detalles_venta = [
            DetalleVenta(
                venta=venta,
                producto=producto,
                cantidad=cantidad,
                subtotal=producto.precio * cantidad,
            )
            for producto, cantidad in productos_objetos
        ]
        try:
            procesar_venta(venta, detalles_venta)
        except ValueError as e:
            messages.error(request, str(e))
            venta.delete()
            return redirect("ventas:crear_venta")
        messages.success(request, "Venta guardada correctamente.")
        # Register cash movement
        try:
            caja_abierta = Caja.objects.get(usuario=request.user, esta_abierta=True)
            monto_ingreso = Decimal('0.00')
            if metodo_pago == 'EF':
                monto_ingreso = venta.total
            elif metodo_pago == 'OT':
                monto_ingreso = venta.mixto_efectivo or Decimal('0.00')
            # No ingreso for TC, TD, TR
            if monto_ingreso > 0:
                MovimientoCaja.objects.create(
                    caja=caja_abierta,
                    tipo='INGRESO',
                    monto=monto_ingreso,
                    descripcion=f"Venta #{venta.id}"
                )
        except Caja.DoesNotExist:
            messages.warning(request, "No se encontró caja abierta para registrar el ingreso. La venta se guardó, pero no se registró en caja.")
        url = reverse("ventas:crear_venta") + f"?venta_id={venta.id}"
        return HttpResponseRedirect(url)
    context = {
        "hoy": now(),
        "metodos_pago": Venta.METODOS_PAGO,
    }
    return render(request, 'ventas/crear_venta.html', context)

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def ticket_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    detalles = venta.detalles.select_related('producto')
    context = {
        "venta": venta,
        "detalles": detalles,
        "hoy": now(),
    }
    return render(request, "ventas/ticket_venta.html", context)

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def get_stock(request, producto_id):                    # (mantiene el mismo nombre)
    """
    Devuelve stock, precio y la URL de la miniatura.
    """
    try:
        producto = Producto.objects.get(id=producto_id)
        # Si el producto no tiene imagen usamos una genérica.
        if producto.imagen:
            imagen_url = producto.imagen.url
        else:
            imagen_url = static('img/no-image.png')
        return JsonResponse({
            'stock': producto.stock,
            'precio': str(producto.precio),
            'imagen_url': imagen_url,
        })
    except Producto.DoesNotExist:
        raise Http404("Producto no encontrado")
    
@grupo_requerido('Encargado', 'Gerente')
def editar_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        form = VentaForm(request.POST, instance=venta)
        formset = DetalleVentaFormSet(request.POST, instance=venta)
        print("Datos POST:", request.POST)
        if form.is_valid() and formset.is_valid():
            detalles = formset.save(commit=False)
            detalles_eliminados = formset.deleted_objects
            print("Detalles enviados:", [(d.producto.nombre, d.cantidad, d.pk) for d in detalles if d.producto])
            print("Detalles eliminados:", [(d.producto.nombre, d.cantidad) for d in detalles_eliminados])
            try:
                with transaction.atomic():
                    # Guardar el formulario principal
                    form.save()
                    # Restaurar stock para detalles eliminados
                    for detalle in detalles_eliminados:
                        detalle.producto.stock += detalle.cantidad
                        detalle.producto.save()
                        print(f"Producto eliminado: {detalle.producto.nombre}, Stock restaurado: {detalle.producto.stock}")
                        detalle.delete()
                    # Procesar y guardar detalles válidos
                    total = Decimal('0.00')
                    for detalle in detalles:
                        if detalle.producto and detalle.cantidad:
                            cantidad_original = 0
                            if detalle.pk:
                                try:
                                    original_detalle = DetalleVenta.objects.get(pk=detalle.pk)
                                    cantidad_original = original_detalle.cantidad
                                except DetalleVenta.DoesNotExist:
                                    print(f"Advertencia: No se encontró detalle con ID {detalle.pk}")
                            detalle.venta = venta
                            detalle.subtotal = Decimal(detalle.producto.precio) * Decimal(detalle.cantidad)
                            detalle.save()
                            detalle.producto.stock = detalle.producto.stock + cantidad_original - detalle.cantidad
                            detalle.producto.save()
                            print(f"Producto: {detalle.producto.nombre}, Cantidad original: {cantidad_original}, Cantidad nueva: {detalle.cantidad}, Stock actualizado: {detalle.producto.stock}")
                            total += detalle.subtotal
                    # Recalcular el total desde los detalles guardados para mayor seguridad
                    total = sum(
                        Decimal(detalle.subtotal)
                        for detalle in DetalleVenta.objects.filter(venta=venta)
                    )
                    venta.total = total
                    venta.save()
                    print(f"Total calculado y guardado: {total}")
                    messages.success(request, "Venta actualizada correctamente.")
                    return redirect('ventas:lista_ventas')
            except Exception as e:
                messages.error(request, f"Error al actualizar la venta: {e}")
        else:
            print("Errores del formulario:", form.errors)
            print("Errores del formset:", formset.errors)
            print("Errores no asociados a campos del formset:", formset.non_form_errors())
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = VentaForm(instance=venta)
        formset = DetalleVentaFormSet(instance=venta)
    productos = Producto.objects.all()
    productos_json = json.dumps({str(p.id): str(p.precio) for p in productos})
    return render(request, 'ventas/editar_venta.html', {
        'form': form,
        'formset': formset,
        'venta': venta,
        'productos_json': productos_json,
    })
    
@grupo_requerido('Encargado', 'Gerente')
def anular_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == "POST":
        try:
            with transaction.atomic():
                revertir_venta(venta, request.user)  # Ya hace todo internamente
                messages.success(request, f"Venta #{venta.id} anulada correctamente.")
        except Exception as e:
            messages.error(request, f"Error al anular la venta: {e}")
        return redirect('ventas:lista_ventas')
    return render(request, 'ventas/anular_venta_confirmacion.html', {'venta': venta})

@grupo_requerido('Encargado', 'Gerente')
def export_venta_pdf(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id)
    detalles = venta.detalles.all()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="venta_{venta_id}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Detalle de Venta #{venta.id}", styles['Title']),
        Spacer(1, 12)
    ]
    info = [
        ['Fecha:', venta.fecha.strftime('%Y-%m-%d %H:%M')],
        ['Método de Pago:', venta.metodo_pago],
        ['Estatus:', "Anulada" if venta.anulada else "Activa"],
    ]
    table_info = Table(info, colWidths=[120, 300])
    table_info.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
    ]))
    elements.append(table_info)
    elements.append(Spacer(1, 20))
    data = [['Producto', 'Cantidad', 'Precio Unitario', 'Subtotal']]
    for d in detalles:
        data.append([
            d.producto.nombre,
            str(d.cantidad),
            f"${d.producto.precio:.2f}",
            f"${d.subtotal:.2f}"
        ])
    data.append(['', '', 'Total:', f"${venta.total:.2f}"])
    table = Table(data, colWidths=[220, 70, 100, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 1), (-1, -2), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

@grupo_requerido('Encargado', 'Gerente')
def export_ventas_excel(request):
    ventas = Venta.objects.all().order_by('-fecha')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    headers = ["ID Venta", "Fecha", "Método Pago", "Estatus", "Producto", "Cantidad", "Precio Unitario", "Subtotal", "Total Venta"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for venta in ventas:
        detalles = venta.detalles.all()
        metodo = dict(Venta.METODOS_PAGO).get(venta.metodo_pago, "Desconocido")
        estatus = "Anulada" if venta.anulada else "Activa"
        for detalle in detalles:
            ws.append([
                venta.id,
                venta.fecha.strftime("%Y-%m-%d %H:%M"),
                metodo,
                estatus,
                detalle.producto.nombre,
                detalle.cantidad,
                detalle.producto.precio,
                detalle.subtotal,
                venta.total,
            ])
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ventas.xlsx"'
    wb.save(response)
    return response

@grupo_requerido('Encargado', 'Gerente')
def reporte_ventas(request):
    hoy = timezone.now().date()
    dias = 7
    ventas_por_dia = (
        Venta.objects.filter(fecha__date__gte=hoy - timezone.timedelta(days=dias))
        .annotate(dia=TruncDate('fecha'))
        .values('dia')
        .annotate(total=Sum('total'))
        .order_by('dia')
    )
    x_dias = [v['dia'].strftime('%d-%b') for v in ventas_por_dia]
    y_totales = [float(Decimal(v['total'])) for v in ventas_por_dia]
    layout_bar = go.Layout(
        title='Ventas últimos 7 días',
        paper_bgcolor='#1f2937',   # gris oscuro más suave
        plot_bgcolor='#1f2937',
        font=dict(color='#a3e635', family='Arial, sans-serif', size=14),
        yaxis=dict(
            title='Total ($)',
            gridcolor='#374151',
            zerolinecolor='#374151',
            tickcolor='#a3e635',
            showline=True,
            linecolor='#a3e635',
        ),
        xaxis=dict(
            gridcolor='#374151',
            tickcolor='#a3e635',
            showline=True,
            linecolor='#a3e635',
        ),
        margin=dict(t=50, b=50, l=70, r=30),
        hovermode='x unified',
        transition={'duration': 500, 'easing': 'cubic-in-out'}
    )
    chart_dias = opy.plot(go.Figure(
        data=[go.Bar(
            x=x_dias,
            y=y_totales,
            marker_color='rgba(163,230,53,0.85)',
            marker_line=dict(width=1.5, color='rgba(72,157,6,1)')
        )],
        layout=layout_bar
    ), auto_open=False, output_type='div')
    pagos = Venta.objects.values('metodo_pago').annotate(total=Sum('total'))
    etiquetas = [dict(Venta.METODOS_PAGO).get(p['metodo_pago'], p['metodo_pago']) for p in pagos]
    cantidades = [float(p['total']) for p in pagos]
    colores_pie = ['#a3e635', '#65a30d', '#ef4444']  # verde lima, verde oscuro, rojo suave
    layout_pie = go.Layout(
        title_text="Ventas por método de pago",
        paper_bgcolor='#1f2937',
        font=dict(color='#a3e635', family='Arial, sans-serif', size=14),
        legend=dict(font=dict(color='#a3e635')),
        margin=dict(t=50, b=50, l=20, r=20),
        hovermode='closest',
        transition={'duration': 500, 'easing': 'cubic-in-out'}
    )
    chart_pagos = opy.plot(go.Figure(
        data=[go.Pie(
            labels=etiquetas,
            values=cantidades,
            hole=0.3,
            marker=dict(colors=colores_pie),
            hoverinfo='label+percent+value'
        )],
        layout=layout_pie
    ), auto_open=False, output_type='div')
    contexto = {
        'chart_dias': chart_dias,
        'chart_pagos': chart_pagos,
        'total_ventas': Venta.objects.aggregate(Sum('total'))['total__sum'] or 0,
        'total_anuladas': Venta.objects.filter(anulada=True).count(),
        'total_activas': Venta.objects.filter(anulada=False).count(),
    }
    return render(request, 'ventas/reporte.html', contexto)

@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    detalles = venta.detalles.select_related('producto')
    # Mapear método de pago al texto legible
    metodo_pago_texto = dict(Venta.METODOS_PAGO).get(venta.metodo_pago, venta.metodo_pago)
    if venta.metodo_pago == 'OT':
        metodo_pago_texto = f"Mixto (Efectivo: ${venta.mixto_efectivo or 0:.2f}, Otros: ${venta.mixto_otros or 0:.2f})"
    # Nuevo: verificar si el usuario tiene los grupos 'Encargado' o 'Gerente'
    user_has_permission = request.user.groups.filter(name__in=['Encargado', 'Gerente']).exists()
    context = {
        'venta': venta,
        'detalles': detalles,
        'metodo_pago_texto': metodo_pago_texto,
        'user_has_permission': user_has_permission,
    }
    return render(request, 'ventas/detalle_venta.html', context)

@require_GET
@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def obtener_info_producto(request):
    producto_id = request.GET.get('producto_id')
    if not producto_id:
        return JsonResponse({'error': 'No se recibió el ID del producto'}, status=400)
    try:
        producto = Producto.objects.get(id=producto_id)
        return JsonResponse({
            'precio': float(producto.precio),
            'stock': producto.stock
        })
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)

@grupo_requerido('Encargado', 'Gerente')
def historial_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    # Obtener todas las entradas del historial ordenadas por fecha descendente
    historial = venta.history.all().order_by('-history_date')
    # Crear lista de tuplas (actual, siguiente, cambios) para comparar
    pares_historial = []
    for i, actual in enumerate(historial):
        siguiente = historial[i + 1] if i + 1 < len(historial) else None
        cambios = get_detailed_changes(actual, siguiente) if siguiente else []
        pares_historial.append((actual, siguiente, cambios))
    context = {
        'venta': venta,
        'pares_historial': pares_historial,
    }
    return render(request, 'ventas/historial_venta.html', context)

def get_detailed_changes(actual, siguiente):
    cambios = actual.diff_against(siguiente).changes
    # Obtener detalles de los productos (detalles) y compararlos
    actual_detalles = {d.pk: d for d in actual.instance.detalles.all()}
    siguiente_detalles = {d.pk: d for d in siguiente.instance.detalles.all()}
    
    # Procesar cambios en detalles
    for pk, detalle in actual_detalles.items():
        producto = detalle.producto
        stock_antes = producto.stock + detalle.cantidad  # Stock antes del cambio
        if pk not in siguiente_detalles:
            # Detalle eliminado
            stock_despues = stock_antes + detalle.cantidad  # Stock aumenta al eliminar
            cambios.append({
                'field': f'detalle_{detalle.producto.nombre}',
                'old': {
                    'cantidad': detalle.cantidad,
                    'precio_unitario': str(detalle.producto.precio),
                    'subtotal': str(detalle.subtotal),
                    'stock': stock_antes,
                    'status': 'eliminado'
                },
                'new': None
            })
        elif detalle != siguiente_detalles[pk]:
            # Detalle modificado
            stock_despues = stock_antes + detalle.cantidad - siguiente_detalles[pk].cantidad  # Ajustar stock
            cambios.append({
                'field': f'detalle_{detalle.producto.nombre}',
                'old': {
                    'cantidad': detalle.cantidad,
                    'precio_unitario': str(detalle.producto.precio),
                    'subtotal': str(detalle.subtotal),
                    'stock': stock_antes
                },
                'new': {
                    'cantidad': siguiente_detalles[pk].cantidad,
                    'precio_unitario': str(siguiente_detalles[pk].producto.precio),
                    'subtotal': str(siguiente_detalles[pk].subtotal),
                    'stock': stock_despues
                }
            })
    
    # Procesar detalles añadidos
    for pk, detalle in siguiente_detalles.items():
        if pk not in actual_detalles:
            producto = detalle.producto
            stock_antes = producto.stock + detalle.cantidad  # Stock antes del cambio
            stock_despues = producto.stock  # Stock después (ya se restó la cantidad)
            cambios.append({
                'field': f'detalle_{detalle.producto.nombre}',
                'old': None,
                'new': {
                    'cantidad': detalle.cantidad,
                    'precio_unitario': str(detalle.producto.precio),
                    'subtotal': str(detalle.subtotal),
                    'stock': stock_despues,
                    'status': 'añadido'
                }
            })
    
    return cambios

@grupo_requerido('Encargado', 'Gerente')
def comparar_historial_venta(request, venta_id, version_a, version_b):
    venta = get_object_or_404(Venta, pk=venta_id)
    v1 = get_object_or_404(venta.history, history_id=version_a)
    v2 = get_object_or_404(venta.history, history_id=version_b)
    
    # Generar los cambios usando get_detailed_changes
    diferencias = get_detailed_changes(v1, v2)
    
    context = {
        'venta': venta,
        'v1': v1,
        'v2': v2,
        'diferencias': diferencias,
    }
    
    return render(request, 'ventas/comparar_venta.html', context)

@require_GET
@grupo_requerido('Cajero', 'Encargado', 'Gerente')
def producto_info(request):
    nombre = request.GET.get('nombre', '').strip()
    if not nombre:
        return JsonResponse([], safe=False)  # Devuelve lista vacía si no hay nombre
    productos = Producto.objects.filter(Q(nombre__icontains=nombre))[:10]
    resultados = [{
        'id': producto.id,
        'nombre': producto.nombre,
        'precio': float(producto.precio),
        'stock': producto.stock,
    } for producto in productos]
    return JsonResponse(resultados, safe=False)