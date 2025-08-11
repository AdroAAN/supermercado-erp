from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .forms import CrearUsuarioForm, EditarUsuarioForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.contrib.auth import views as auth_views

# Reutiliza el decorador de permisos por grupo (debes tenerlo definido en un archivo utilitario)
def grupo_requerido(*nombres_grupos):
    def decorador(view_func):
        @login_required
        def _wrapped_view(request, *args, **kwargs):
            if request.user.groups.filter(name__in=nombres_grupos).exists() or request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            messages.error(request, "No tienes permisos para acceder a esta sección.")
            return redirect('reportes:dashboard')  # Ajusta si tu ruta del dashboard es diferente
        return _wrapped_view
    return decorador

@grupo_requerido('Gerente')
def crear_usuario(request):
    if request.method == 'POST':
        form = CrearUsuarioForm(request.POST)
        grupo_id = request.POST.get('grupo')  # Nombre del grupo seleccionado
        if form.is_valid():
            usuario = form.save()
            if grupo_id:
                try:
                    grupo = Group.objects.get(id=grupo_id)
                    usuario.groups.add(grupo)
                except Group.DoesNotExist:
                    messages.error(request, f"El grupo con ID '{grupo_id}' no existe.")
            messages.success(request, 'Usuario creado correctamente.')
            return redirect('accounts:listar_usuarios')
    else:
        form = CrearUsuarioForm()

    grupos = Group.objects.all()
    return render(request, 'accounts/crear_usuario.html', {'form': form, 'grupos': grupos})

@grupo_requerido('Gerente')
def editar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.save()
            # Limpia grupos y agrega el grupo seleccionado desde el formulario
            usuario.groups.clear()
            grupo = form.cleaned_data.get('grupo')
            if grupo:
                usuario.groups.add(grupo)
            messages.success(request, 'Usuario actualizado.')
            return redirect('accounts:listar_usuarios')
    else:
        form = EditarUsuarioForm(instance=usuario)

    grupos = Group.objects.all()
    return render(request, 'accounts/editar_usuario.html', {
        'form': form,
        'usuario': usuario,
        'grupos': grupos,
        'grupo_actual': usuario.groups.first()
    })

@grupo_requerido('Gerente')
def desactivar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        usuario.is_active = False
        usuario.save()
        messages.success(request, 'Usuario desactivado. Ya no podrá iniciar sesión.')
        return redirect('accounts:listar_usuarios')
    return render(request, 'accounts/desactivar_usuario.html', {'usuario': usuario})

@grupo_requerido('Gerente')
def reactivar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        usuario.is_active = True
        usuario.save()
        messages.success(request, 'Usuario reactivado correctamente.')
        return redirect('accounts:listar_usuarios')
    return render(request, 'accounts/reactivar_usuario.html', {'usuario': usuario})

@grupo_requerido('Gerente')
def listar_usuarios(request):
    estado = request.GET.get('estado', '')  # leer parámetro GET 'estado'
    
    if estado == 'activos':
        usuarios = User.objects.filter(is_active=True)
    elif estado == 'inactivos':
        usuarios = User.objects.filter(is_active=False)
    else:
        usuarios = User.objects.all()
    
    return render(request, 'accounts/lista_usuarios.html', {'usuarios': usuarios})

@grupo_requerido('Gerente')
def exportar_usuarios_excel(request):
    usuarios = User.objects.all().order_by('username')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ["ID", "Usuario", "Email", "Roles", "Activo", "Es staff"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for user in usuarios:
        roles = ", ".join(g.name for g in user.groups.all()) or "Sin roles"
        fila = [
            user.id,
            user.username,
            user.email,
            roles,
            "Sí" if user.is_active else "No",
            "Sí" if user.is_staff else "No"
        ]
        ws.append(fila)

    # Ajustar anchos
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
    wb.save(response)
    return response

class CustomPasswordResetView(auth_views.PasswordResetView):
    template_name = 'registration/password_reset_form.html'
    email_template_name = 'registration/password_reset_email.html'
    subject_template_name = 'registration/password_reset_subject.txt'
    success_url = '/accounts/password_reset/done/'

class CustomPasswordResetDoneView(auth_views.PasswordResetDoneView):
    template_name = 'registration/password_reset_done.html'

class CustomPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'registration/password_reset_confirm.html'
    success_url = '/accounts/reset/done/'

class CustomPasswordResetCompleteView(auth_views.PasswordResetCompleteView):
    template_name = 'registration/password_reset_complete.html'